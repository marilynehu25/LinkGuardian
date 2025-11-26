# librairie Flask

# --- Librairies Python standards ---
from datetime import datetime
from io import BytesIO
from math import ceil
from urllib.parse import urlparse

import pandas as pd
import requests
from flask import (
    Blueprint,
    abort,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, func
from sqlalchemy.orm import joinedload

# à partir du fichier python database.py
from database import db
from models import Source, Tag, TaskRecord, User, UserAccess, Website
from services.access_service import user_can_access_data
from services.api_babbar import fetch_url_data
from services.check_service import (
    check_link_presence_and_follow_status,
    perform_check_status,
)
from services.stats_service import save_stats_snapshot
from services.utils_service import check_anchor_presence, couleur_aleatoire_unique
from tasks import check_all_user_sites

sites_routes = Blueprint("sites_routes", __name__)


def extract_domain(url):
    """Récupère le domaine principal d'une URL."""
    try:
        parsed = urlparse(url)
        return parsed.netloc.lower().replace("www.", "")
    except Exception:
        return ""


@sites_routes.route("/add_site", methods=["POST"])
def add_site():
    url = request.form.get("url", "").strip()
    tag = request.form.get("tag", "").strip().lower()
    link_to_check = request.form.get("link_to_check", "").strip()
    anchor_text = request.form.get("anchor_text", "").strip()
    source_plateforme = request.form.get("source_plateforme", "").strip()

    # ✅ VALIDATION AMÉLIORÉE - Vérifier TOUS les champs obligatoires
    if not url or not tag or not link_to_check:
        flash(
            "⚠️ Veuillez remplir tous les champs obligatoires (URL, Tag, Lien à vérifier).",
            "warning",
        )

        # Si appel HTMX → ne rien faire (pas de rendu de tableau)
        if request.headers.get("HX-Request"):
            # Retourner un message d'erreur au lieu du tableau
            return (
                """
                <div class="bg-yellow-500/10 border border-yellow-500 text-yellow-500 px-4 py-3 rounded-lg mb-4">
                    ⚠️ Veuillez remplir tous les champs obligatoires
                </div>
            """,
                400,
            )

        return redirect(request.referrer or url_for("main_routes.index"))

    # ✅ VALIDATION DES URLs
    if not url.startswith(("http://", "https://")):
        flash("⚠️ L'URL doit commencer par http:// ou https://", "warning")
        if request.headers.get("HX-Request"):
            return (
                """
                <div class="bg-yellow-500/10 border border-yellow-500 text-yellow-500 px-4 py-3 rounded-lg mb-4">
                    ⚠️ L'URL doit commencer par http:// ou https://
                </div>
            """,
                400,
            )
        return redirect(request.referrer or url_for("main_routes.index"))

    if not link_to_check.startswith(("http://", "https://")):
        flash("⚠️ Le lien à vérifier doit commencer par http:// ou https://", "warning")
        if request.headers.get("HX-Request"):
            return (
                """
                <div class="bg-yellow-500/10 border border-yellow-500 text-yellow-500 px-4 py-3 rounded-lg mb-4">
                    ⚠️ Le lien à vérifier doit commencer par http:// ou https://
                </div>
            """,
                400,
            )
        return redirect(request.referrer or url_for("main_routes.index"))

    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        html_content = response.text

        link_present, follow_status = check_link_presence_and_follow_status(
            html_content, link_to_check
        )
        anchor_present = check_anchor_presence(html_content, anchor_text)

        new_site = Website(
            url=url,
            domains=extract_domain(url),
            tag=tag,
            link_to_check=link_to_check,
            anchor_text=anchor_text,
            source_plateforme=source_plateforme,
            user_id=current_user.id,  # ✅ CORRECTION : utiliser current_user.id au lieu de current_user.email
            link_status="Lien présent" if link_present else "Lien absent",
            anchor_status="Ancre présente" if anchor_present else "Ancre absente",
            link_follow_status=follow_status if link_present else None,
            first_checked=datetime.now(),
            last_checked=datetime.now(),
        )

        db.session.add(new_site)
        db.session.commit()

        # ✅ Ces deux lignes mettent à jour la ligne dans la base
        perform_check_status(new_site.id)
        fetch_url_data(new_site.url, async_mode=False)

        # ✅ On recharge depuis la DB pour avoir les dernières valeurs
        db.session.refresh(new_site)

        flash("✅ Site ajouté et vérifié avec succès !", "success")

        # Si appel HTMX → renvoie le tableau actualisé
        if request.headers.get("HX-Request"):
            # ✅ Récupérer la requête filtrée avec pagination
            query = Website.query.filter_by(user_id=current_user.id).order_by(
                Website.id.desc()
            )

            page = 1
            per_page = 10
            pagination = query.paginate(page=page, per_page=per_page, error_out=False)

            return render_template(
                "backlinks/_table.html",
                backlinks=pagination.items,
                current_page=pagination.page,
                total_pages=pagination.pages or 1,
            )

    except requests.Timeout:
        flash("⏱️ Timeout : Le site met trop de temps à répondre", "danger")
        if request.headers.get("HX-Request"):
            return (
                """
                <div class="bg-red-500/10 border border-red-500 text-red-500 px-4 py-3 rounded-lg mb-4">
                    ⏱️ Le site met trop de temps à répondre
                </div>
            """,
                500,
            )

    except requests.RequestException as e:
        flash(f"❌ Erreur lors de la vérification de l'URL : {e}", "danger")
        if request.headers.get("HX-Request"):
            return (
                f"""
                <div class="bg-red-500/10 border border-red-500 text-red-500 px-4 py-3 rounded-lg mb-4">
                    ❌ Erreur : {str(e)}
                </div>
            """,
                500,
            )

    db.session.refresh(new_site)
    save_stats_snapshot(current_user.id)
    flash("✅ Site ajouté et vérifié avec succès !", "success")

    return redirect(request.referrer or url_for("main_routes.index"))


# cette fonction permet à l'utilisateur de supprimer un site de la base de données en fonction de son identifiant,
# puis elle redirige l'utilisateur vers la page d'accueil.
@sites_routes.route("/delete_site/<int:site_id>", methods=["POST"])
def delete_site(site_id):
    site_to_delete = Website.query.get(site_id)
    if not site_to_delete:
        print("❌ Site non trouvé :", site_id)
        return "Site non trouvé", 404

    if not user_can_access_data(current_user.id, site_to_delete.user_id):
        abort(403)

    try:
        print(f"🗑️ Suppression du site ID {site_id} → {site_to_delete.url}")

        # Supprime les doublons liés
        duplicates = Website.query.filter(
            and_(
                Website.url == site_to_delete.url,
                Website.link_to_check == site_to_delete.link_to_check,
                Website.id != site_to_delete.id,
            )
        ).all()

        for duplicate in duplicates:
            print(f"  ↳ Duplicate supprimé : {duplicate.id}")
            db.session.delete(duplicate)

        db.session.delete(site_to_delete)
        db.session.commit()
        save_stats_snapshot(current_user.id)
        print("✅ Suppression réussie")
        return "", 204  # No Content

    except Exception as e:
        db.session.rollback()
        print("❌ Erreur lors de la suppression :", e)
        return "Erreur lors de la suppression", 500


# Une fonction est conçue pour déclencher la vérification du statut du lien et du texte d'ancre, ainsi que la mise à jour des données Babbar pour un site spécifié.
# Après avoir effectué ces opérations, elle sauvegarde les changements dans la base de données et redirige l'utilisateur vers la page d'accueil.
@sites_routes.route("/check_status/<int:site_id>", methods=["GET", "POST"])
def check_status(site_id):
    """Vérifie et met à jour le statut d'un site"""
    site = Website.query.get_or_404(site_id)

    if not user_can_access_data(current_user.id, site.user_id):
        abort(403)

    try:
        # Effectuer les vérifications et mises à jour
        perform_check_status(site.id)
        babbar_data = fetch_url_data(site.url, async_mode=False)

        if babbar_data:
            site.page_value = babbar_data.get("pageValue")
            site.page_trust = babbar_data.get("pageTrust")
            site.bas = babbar_data.get("babbarAuthorityScore")
            site.backlinks_external = babbar_data.get("backlinksExternal")
            site.num_outlinks_ext = babbar_data.get("numOutLinksExt")

        # Mettre à jour la date de vérification
        site.last_checked = datetime.now()
        if site.first_checked is None:
            site.first_checked = datetime.now()

        db.session.commit()

        # ✅ Recharger depuis la DB pour avoir les dernières valeurs
        site = Website.query.get(site.id)

        print(f"✅ Site vérifié : {site.url}")
        print(f"   - Status HTTP: {site.status_code}")
        print(f"   - Link status: {site.link_status}")
        print(f"   - Follow status: {site.link_follow_status}")
        print(f"   - Google index: {site.google_index_status}")

    except Exception:
        db.session.rollback()
        import traceback

        print("❌ ERREUR RÉELLE COMPLÈTE ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓")
        traceback.print_exc()

    save_stats_snapshot(current_user.id)

    # Retourner la ligne mise à jour (HTMX) ou rediriger
    if request.headers.get("HX-Request"):
        return render_template("backlinks/_row.html", backlink=site)
    else:
        return redirect(url_for("main_routes.index"))


# cette fonction sert à Supprimer tous les sites de la base de données
@sites_routes.route("/delete_all_sites", methods=["POST"])
def delete_all_sites():
    from celery_app import celery

    records = TaskRecord.query.filter_by(user_id=current_user.id).all()
    for r in records:
        celery.control.revoke(r.task_id, terminate=True)

    TaskRecord.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()

    Website.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()
    save_stats_snapshot(current_user.id)
    flash("✅ Tous les sites ont été supprimés avec succès.", "success")
    return redirect(url_for("backlinks_routes.backlinks_list"))


# Route pour forcer la vérification manuelle de TOUS les sites de l'utilisateur
@sites_routes.route("/check_all_sites", methods=["POST"])
@login_required
def check_all_sites():
    print("=" * 60)
    print("🔍 [DEBUG] check_all_sites() appelée")
    print(f"🔍 [DEBUG] User ID: {current_user.id}")
    print(f"🔍 [DEBUG] Request method: {request.method}")
    print(f"🔍 [DEBUG] Headers: {dict(request.headers)}")
    print("=" * 60)

    try:
        print("🔍 [DEBUG] Tentative d'import de check_all_user_sites...")

        print("✅ [DEBUG] Import réussi")

        print("🔍 [DEBUG] Tentative de lancement de la tâche...")
        result = check_all_user_sites.delay(current_user.id)
        print(f"✅ [DEBUG] Tâche lancée avec ID: {result.id}")

        # 🟢 ENREGISTRER LE TASK ID POUR LA PURGE PERSONNALISÉE
        record = TaskRecord(task_id=result.id, user_id=current_user.id)
        db.session.add(record)
        db.session.commit()
        print(f"🟢 [DEBUG] TaskRecord sauvegardé : {result.id}")

        flash("🚀 Vérification globale lancée en arrière-plan !", "success")

    except Exception as e:
        print(f"❌ [DEBUG] Erreur générale: {type(e).__name__}: {e}")
        import traceback

        traceback.print_exc()
        flash(f"❌ Erreur : {e}", "danger")

    print("🔍 [DEBUG] Fin de la fonction, redirection...")
    print("=" * 60)

    return redirect(url_for("backlinks_routes.backlinks_list"))


@sites_routes.route("/import", methods=["GET", "POST"])
def import_data():
    if request.method == "POST":
        file = request.files.get("file")
        if not file:
            flash("Aucun fichier sélectionné", "error")
            return redirect(request.referrer)

        try:
            df = pd.read_excel(file)
            df.columns = [col.lower().strip() for col in df.columns]

            # 🔥 Déduplication immédiate
            df = df.drop_duplicates(subset=["url", "link_to_check"], keep="last")

            # 🔥 Prétraitement
            df["url"] = df["url"].astype(str).str.strip()
            df["tag"] = df["tag"].astype(str).str.lower().str.strip()
            df["plateforme"] = df["plateforme"].astype(str).str.lower().str.strip()
            df["link_to_check"] = df["link_to_check"].astype(str).str.strip()
            df["anchor_text"] = df["anchor_text"].astype(str).str.strip()

            # 🔥 Précharger Tags & Sources existants (ultra rapide)
            existing_tags = {t.valeur.lower(): t for t in Tag.query.all()}
            existing_sources = {s.nom.lower(): s for s in Source.query.all()}

            # 🔥 Précharger les sites existants de l’utilisateur
            existing_sites = Website.query.filter_by(user_id=current_user.id).all()
            lookup = {(s.url, s.link_to_check): s for s in existing_sites}

            new_sites = []
            updated_sites = []
            websites_to_check = []

            # 🔥 Boucle principale
            for _, row in df.iterrows():
                url = row["url"]
                if not url:
                    continue

                key = (url, row["link_to_check"])
                domain = extract_domain(url)

                tag_value = row["tag"] if row["tag"] else None
                source_value = row["plateforme"] if row["plateforme"] else None

                # -------------------------
                # 🏷️ GESTION AUTO DES TAGS
                # -------------------------
                if tag_value:
                    tag_key = tag_value.lower()

                    if tag_key not in existing_tags:
                        new_tag = Tag(
                            valeur=tag_key, couleur=couleur_aleatoire_unique()
                        )
                        db.session.add(new_tag)
                        existing_tags[tag_key] = new_tag

                # ----------------------------------
                # 🏭 GESTION AUTO DES SOURCES
                # ----------------------------------
                if source_value:
                    source_key = source_value.lower()

                    if source_key not in existing_sources:
                        new_source = Source(nom=source_key)
                        db.session.add(new_source)
                        existing_sources[source_key] = new_source

                # -------------------------
                # 🔄 UPDATE d’un site existant
                # -------------------------
                if key in lookup:
                    site = lookup[key]

                    site.tag = tag_value or site.tag
                    site.domains = domain or site.domains
                    site.source_plateforme = source_value or site.source_plateforme
                    site.anchor_text = row["anchor_text"] or site.anchor_text

                    updated_sites.append(site)

                # -------------------------
                # 🆕 INSERT d’un nouveau site
                # -------------------------
                else:
                    site = Website(
                        url=url,
                        domains=domain,
                        tag=tag_value,
                        link_to_check=row["link_to_check"],
                        anchor_text=row["anchor_text"],
                        source_plateforme=source_value,
                        user_id=current_user.id,
                        first_checked=datetime.now(),
                    )
                    new_sites.append(site)

                websites_to_check.append(site)

            # 🔥 Commit global (sites + nouveaux tags/sources)
            db.session.add_all(new_sites)
            db.session.commit()

            # -------------------------
            # 🚀 Envoi des tâches Celery
            # -------------------------
            from tasks import check_single_site
            
            task_records = []
            for site in websites_to_check:
                task = check_single_site.apply_async(
                    args=[site.id], queue="standard", priority=3
                )
                task_records.append(
                    TaskRecord(task_id=task.id, user_id=current_user.id)
                )

            db.session.add_all(task_records)
            db.session.commit()

            flash(
                "Import lancé 🚀 Les vérifications se font en arrière-plan.", "success"
            )

        except Exception as e:
            db.session.rollback()
            print("Erreur import:", e)
            flash("Erreur lors de l'import.", "error")

        return redirect(url_for("backlinks_routes.backlinks_list"))

    # GET → Affichage liste
    websites = Website.query.filter_by(user_id=current_user.id).all()
    stats = calculate_stats(current_user.id)
    sources = Source.query.all()

    save_stats_snapshot(current_user.id)

    return render_template(
        "backlinks/list.html",
        backlinks=websites,
        stats=stats,
        current_page=1,
        total_pages=1,
        sort="created",
        order="desc",
        sources=sources,
    )


def calculate_stats(user_id):
    websites = Website.query.filter_by(user_id=user_id).all()
    total = len(websites)
    follow_count = sum(1 for w in websites if w.link_follow_status == "follow")
    indexed_count = sum(1 for w in websites if w.google_index_status == "indexed")
    stats = {
        "total": total,
        "follow": follow_count,
        "follow_percentage": f"{(follow_count / total * 100) if total > 0 else 0:.1f}",
        "indexed": indexed_count,
        "indexed_percentage": f"{(indexed_count / total * 100) if total > 0 else 0:.1f}",
    }
    return stats


if False:

    @sites_routes.route("/import", methods=["GET", "POST"])
    def import_data():
        if request.method == "POST":
            file = request.files.get("file")
            if not file:
                flash("Aucun fichier sélectionné", "error")
                return redirect(request.referrer)

            try:
                df = pd.read_excel(file)
                df.columns = [col.lower().strip() for col in df.columns]

                websites_to_check = []

                for _, row in df.iterrows():
                    url = str(row.get("url", "")).strip()
                    if not url:
                        continue

                    tag = str(row.get("tag", "")).lower().strip()
                    domain = extract_domain(url)
                    source_plateforme = str(row.get("plateforme", "")).strip()
                    link_to_check = str(row.get("link_to_check", "")).strip()
                    anchor_text = str(row.get("anchor_text", "")).strip()

                    site = Website.query.filter_by(
                        url=url, link_to_check=link_to_check, user_id=current_user.id
                    ).first()

                    if site:
                        site.tag = tag or site.tag
                        site.domains = domain or site.domains
                        site.source_plateforme = (
                            source_plateforme or site.source_plateforme
                        )
                        site.anchor_text = anchor_text or site.anchor_text
                    else:
                        site = Website(
                            url=url,
                            domains=domain,
                            tag=tag,
                            link_to_check=link_to_check,
                            anchor_text=anchor_text,
                            source_plateforme=source_plateforme,
                            user_id=current_user.id,
                            first_checked=datetime.now(),
                        )
                        db.session.add(site)

                    websites_to_check.append(site)

                db.session.commit()

                # Lancer Celery
                from tasks import check_single_site

                task_records = []  # Buffer

                for site in websites_to_check:
                    task = check_single_site.apply_async(
                        args=[site.id], queue="standard", priority=3
                    )

                    task_records.append(
                        TaskRecord(task_id=task.id, user_id=current_user.id)
                    )

                # Ajout des enregistrements en une seule fois
                db.session.add_all(task_records)
                db.session.commit()

                flash(
                    "Import lancé 🚀 Les vérifications se font en arrière-plan.",
                    "success",
                )

            except Exception as e:
                db.session.rollback()
                print("Erreur import:", e)
                flash("Erreur lors de l'import.", "error")

            return redirect(url_for("backlinks_routes.backlinks_list"))

        # GET → liste normale
        websites = Website.query.filter_by(user_id=current_user.id).all()
        stats = calculate_stats(current_user.id)
        sources = Source.query.all()

        return render_template(
            "backlinks/list.html",
            backlinks=websites,
            stats=stats,
            current_page=1,
            total_pages=1,
            sort="created",
            order="desc",
            sources=sources,
        )


# bouton pour exporter les données en CSV
@sites_routes.route("/export_data", methods=["GET"])
@login_required
def export_data():
    """Exporte la liste des sites en Excel"""

    # 🔒 Étape 1 : Récupérer user_id depuis l'URL et déterminer pour quel utilisateur on exporte
    user_id = request.args.get("user_id", type=int)
    target_user_id = user_id or current_user.id

    # 🔒 Étape 2 : Vérifier le droit d'accès
    if not user_can_access_data(current_user.id, target_user_id):
        abort(403)

    # Créer un workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlinks"

    # Style pour l'en-tête
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)

    # En-têtes
    headers = [
        "URL",
        "Tag",
        "Plateforme",
        "link_to_check",
        "link_status",
        "anchor_text",
        "anchor_status",
        "link_follow_status",
        "google_index_status",
        "page_value",
        "page_trust",
        "bas",
        "backlinks_external",
        "num_outlinks_ext",
        "last_checked",
    ]
    ws.append(headers)

    # Appliquer le style à l'en-tête
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ✅ Étape 3 : Récupération sécurisée des données
    websites = Website.query.filter_by(user_id=target_user_id).all()

    # Ajouter les données
    for site in websites:
        row = [
            site.url or "",
            site.tag or "",
            site.source_plateforme or "",
            site.link_to_check or "",
            site.link_status or "",
            site.anchor_text or "",
            site.anchor_status or "",
            site.link_follow_status or "",
            site.google_index_status or "",
            site.page_value or "",
            site.page_trust or "",
            site.bas or "",
            site.backlinks_external or "",
            site.num_outlinks_ext or "",
            site.last_checked or "",
        ]
        ws.append(row)

    # Ajuster la largeur des colonnes
    ws.column_dimensions["A"].width = 50  # URL
    ws.column_dimensions["B"].width = 15  # Tag
    ws.column_dimensions["C"].width = 15  # Plateforme source
    ws.column_dimensions["D"].width = 50  # Lien à vérifier
    ws.column_dimensions["E"].width = 30  # Texte d'ancre

    # Sauvegarder dans un buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nom du fichier avec date
    filename = f"LinkGuardian_backlinks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# -------------------------------------------------------------------------------------------------------------------------------------------------
#                                                DONNÉES PARTAGÉES
# -------------------------------------------------------------------------------------------------------------------------------------------------


def get_filtered_query_for_owner(owner_id):
    """Construit une requête filtrée pour un utilisateur dont on consulte les données partagées."""

    query = Website.query.filter_by(user_id=owner_id)

    # -------- Filtres TAG & SOURCE --------
    filter_tag = request.args.get("tag", "").strip()
    filter_source = request.args.get("source", "").strip()

    if filter_tag:
        query = query.filter(func.lower(Website.tag) == filter_tag.lower())

    if filter_source:
        query = query.filter(
            func.lower(Website.source_plateforme) == filter_source.lower()
        )

    # -------- Recherche textuelle --------
    q = request.args.get("q", "").strip()
    if q:
        query = query.filter(
            (Website.url.ilike(f"%{q}%")) | (Website.anchor_text.ilike(f"%{q}%"))
        )

    # -------- Follow / NoFollow --------
    follow = request.args.get("follow", "all")
    if follow == "true":
        query = query.filter(Website.link_follow_status == "follow")
    elif follow == "false":
        query = query.filter(Website.link_follow_status == "nofollow")

    # -------- Indexation --------
    indexed = request.args.get("indexed", "all")
    if indexed == "true":
        query = query.filter(Website.google_index_status == "Indexé !")
    elif indexed == "false":
        query = query.filter(Website.google_index_status != "Indexé !")

    # -------- TRI --------
    sort = request.args.get("sort", "created")
    order = request.args.get("order", "desc")

    if sort == "page_value":
        order_by = (
            Website.page_value.desc() if order == "desc" else Website.page_value.asc()
        )
    elif sort == "page_trust":
        order_by = (
            Website.page_trust.desc() if order == "desc" else Website.page_trust.asc()
        )
    elif sort == "domain":
        order_by = Website.url.desc() if order == "desc" else Website.url.asc()
    else:
        order_by = Website.id.desc() if order == "desc" else Website.id.asc()

    return query.order_by(order_by)


@sites_routes.route("/shared_data", methods=["GET"])
@login_required
def shared_data():
    shared_with_me = (
        UserAccess.query.options(joinedload(UserAccess.owner))
        .filter_by(grantee_id=current_user.id)
        .all()
    )

    selected_owner_id = request.args.get("owner_id", type=int)
    backlinks = []
    selected_owner = None
    current_page = 1
    total_pages = 1
    stats = None

    filters = {
        "q": request.args.get("q", ""),
        "tag": request.args.get("tag", ""),
        "source": request.args.get("source", ""),
        "follow": request.args.get("follow", "all"),
        "indexed": request.args.get("indexed", "all"),
        "sort": request.args.get("sort", "created"),
        "order": request.args.get("order", "desc"),
    }

    pagination_base_url = None  # ✅ important : valeur par défaut

    if selected_owner_id:
        selected_owner = User.query.get_or_404(selected_owner_id)

        if not user_can_access_data(current_user.id, selected_owner.id):
            abort(403)

        per_page = 10
        page = request.args.get("page", 1, type=int)

        query = get_filtered_query_for_owner(selected_owner.id)

        total_items = query.count()
        total_pages = ceil(total_items / per_page) if total_items else 1

        backlinks = query.offset((page - 1) * per_page).limit(per_page).all()
        current_page = page

        # ----- STATISTIQUES FILTRÉES -----
        stats_query = get_filtered_query_for_owner(selected_owner.id).order_by(None)

        total = stats_query.count()

        if total:
            follow_count = stats_query.filter(
                Website.link_follow_status == "follow"
            ).count()
            indexed_count = stats_query.filter(
                Website.google_index_status == "Indexé !"
            ).count()

            avg_value = (
                stats_query.with_entities(func.avg(Website.page_value)).scalar() or 0
            )
            avg_trust = (
                stats_query.with_entities(func.avg(Website.page_trust)).scalar() or 0
            )

            avg_quality = round((float(avg_trust) * 0.6) + (float(avg_value) * 0.4), 1)
        else:
            follow_count = indexed_count = avg_value = avg_trust = avg_quality = 0

        stats = {
            "total": total,
            "follow": follow_count,
            "follow_percentage": f"{(follow_count / total * 100) if total else 0:.1f}",
            "indexed": indexed_count,
            "indexed_percentage": f"{(indexed_count / total * 100) if total else 0:.1f}",
            "avg_quality": f"{avg_quality:.1f}",
            "avg_value": f"{avg_value:.1f}",
            "avg_trust": f"{avg_trust:.1f}",
        }

        # ----- ⚡ CRÉATION DU BASE_URL UNIQUEMENT SI OWNER -----
        pagination_base_url = url_for(
            "sites_routes.shared_data_table_partial",
            owner_id=selected_owner.id,
            q=filters["q"],
            tag=filters["tag"],
            source=filters["source"],
            follow=filters["follow"],
            indexed=filters["indexed"],
            sort=filters["sort"],
            order=filters["order"],
        )

    return render_template(
        "shared/shared_data.html",
        shared_with_me=shared_with_me,
        selected_owner=selected_owner,
        backlinks=backlinks,
        current_page=current_page,
        total_pages=total_pages,
        stats=stats,
        filters=filters,
        tags=Tag.query.all(),
        sources=Source.query.all(),
        pagination_base_url=pagination_base_url,
    )


@sites_routes.route("/shared_data/table", methods=["GET"])
@login_required
def shared_data_table_partial():
    """Partial HTMX - seulement le tableau pour les données partagées"""

    # ⚡ Si ce n’est pas un appel HTMX, redirige vers la page complète
    if not request.headers.get("HX-Request"):
        page = request.args.get("page", 1, type=int)
        owner_id = request.args.get("owner_id", type=int)
        return redirect(
            url_for("sites_routes.shared_data", owner_id=owner_id, page=page)
        )

    owner_id = request.args.get("owner_id", type=int)
    if not owner_id:
        abort(400)

    selected_owner = User.query.get_or_404(owner_id)

    # Vérification d’accès
    if not user_can_access_data(current_user.id, selected_owner.id):
        abort(403)

    per_page = 10
    page = request.args.get("page", 1, type=int)

    query = get_filtered_query_for_owner(selected_owner.id)

    total_items = query.count()
    total_pages = ceil(total_items / per_page) if total_items > 0 else 1

    backlinks = query.offset((page - 1) * per_page).limit(per_page).all()

    # Calcul qualité
    for site in backlinks:
        trust = float(site.page_trust or 0)
        value = float(site.page_value or 0)
        site.quality = round((trust * 0.6) + (value * 0.4), 1) if trust or value else 0

    base_url = url_for(
        "sites_routes.shared_data_table_partial",
        owner_id=selected_owner.id,
        q=request.args.get("q", ""),
        tag=request.args.get("tag", ""),
        source=request.args.get("source", ""),
        follow=request.args.get("follow", "all"),
        indexed=request.args.get("indexed", "all"),
        sort=request.args.get("sort", "created"),
        order=request.args.get("order", "desc"),
    )

    return render_template(
        "backlinks/_table.html",
        backlinks=backlinks,
        current_page=page,
        total_pages=total_pages,
        selected_owner=selected_owner,
        pagination_base_url=base_url,
    )
